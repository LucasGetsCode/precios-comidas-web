# import descargar_precios
import openpyxl

path = "articulos.xlsx"
inicio = 1

def obtener(): 

    def links_super(hoja, fila, columna):
        links_string = hoja.cell(row=fila, column=columna).value
        links_array = links_string.split(",") if links_string else []
        return links_array


    wb = openpyxl.load_workbook(path)
    hoja = wb["Hoja1"]
    cant_filas = hoja.max_row

    links = {}

    for fila in range(inicio, cant_filas):
        articulo = hoja.cell(row=fila, column=1).value.lower()
        
        links_coto = links_super(hoja,fila,2)
        links_dia = links_super(hoja,fila,3)
        links_carrefour = links_super(hoja,fila,4)

        if links[articulo]: print("Este artículo se encuentra repetido")
        links[articulo] = {
            "nombre" : articulo,
            "coto" : links_coto,
            "dia" : links_dia,
            "carrefour" : links_carrefour,
            "fila" : fila
        }

    return links

